import fitz  # PyMuPDF
import os
import re
import time
import string
import pyautogui
from dateutil import parser
from openpyxl import Workbook, load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

class CampaignFinanceDownloader:
    def __init__(self):
        self.script_directory = os.path.dirname(os.path.abspath(__file__))
        self.download_folder = os.path.join(self.script_directory, "Downloaded_Files")
        self.excel_file_path = os.path.join(self.script_directory, "downloaded_files.xlsx")
        self.setup_directories()
        self.setup_driver()

    def setup_directories(self):
        if not os.path.exists(self.download_folder):
            os.makedirs(self.download_folder)
            print(f"📁 Created download folder: {self.download_folder}")

        if not os.path.exists(self.excel_file_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["ID", "First Name", "Last Name", "File Path", "Report Type", "Submit Date",
                       "Report ID", "Upload Report Id", "Campaign Year", "Date Created", "Candidate ID"])
            wb.save(self.excel_file_path)
            print(f"📄 Created Excel log: {self.excel_file_path}")

    def setup_driver(self):
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")
        prefs = {
            "download.prompt_for_download": True,
            "download.directory_upgrade": True,
            "safebrowsing.enabled": True
        }
        options.add_experimental_option("prefs", prefs)
        self.driver = webdriver.Chrome(options=options)

    def handle_save_as_dialog(self, file_path):
        time.sleep(5)
        pyautogui.write(file_path)
        time.sleep(2)
        pyautogui.press("enter")
        time.sleep(2)

    def download_reports(self):
        for letter in string.ascii_uppercase:
        # for letter in string.ascii_uppercase[string.ascii_uppercase.index("A"):string.ascii_uppercase.index("Z") + 1]:

        # for letter in string.ascii_uppercase[string.ascii_uppercase.index("T"):]:
            print(f"\n🔍 Processing names starting with: {letter}")
            try:
                self.driver.get("https://cohweb.houstontx.gov/CampaignFinanceWeb/CFRwebsiteSimpleSearch.aspx")
                time.sleep(3)

                input_field = self.driver.find_element(By.XPATH, "//input[contains(@name, 'txtFirstName_coh')]")
                input_field.clear()
                input_field.send_keys(letter)
                time.sleep(2)

                search_button = self.driver.find_element(By.XPATH, "//input[contains(@name, 'btnSearch_coh')]")
                search_button.click()
                time.sleep(10)

                table = self.driver.find_element(By.ID, "ctl00_ContentPlaceHolder1_grdCandidate")
                rows = table.find_elements(By.TAG_NAME, "tr")

                for index, row in enumerate(rows[1:], start=1):
                    cols = row.find_elements(By.TAG_NAME, "td")
                    if not cols:
                        print(f"Row {index} is empty, skipping.")
                        continue

                    link_elements = cols[0].find_elements(By.TAG_NAME, "a")
                    if not link_elements:
                        print(f"Row {index} has no PDF link, skipping.")
                        continue

                    row_data = [col.text.strip() for col in cols]
                    if len(row_data) < 14:
                        print(f"Skipping row {index} due to insufficient data.")
                        continue

                    id, first_name, last_name = row_data[1], row_data[2], row_data[3]
                    report_type, date_submit = row_data[4], row_data[5]
                    report_id, upload_report_type = row_data[6], row_data[7]
                    campaign_year, date_created, candidate_id = row_data[8], row_data[12], row_data[13]

                    new_pdf_filename = f"{first_name}_{last_name}_{id}_{candidate_id}.pdf"
                    new_pdf_path = os.path.join(self.download_folder, new_pdf_filename)

                    link_elements[0].click()
                    print(f"🔄 Downloading {first_name} {last_name}'s file...")
                    time.sleep(2)
                    self.handle_save_as_dialog(new_pdf_path)

                    wb = load_workbook(self.excel_file_path)
                    ws = wb.active
                    ws.append([id, first_name, last_name, new_pdf_path, report_type, date_submit,
                               report_id, upload_report_type, campaign_year, date_created, candidate_id])
                    wb.save(self.excel_file_path)

            except Exception as e:
                print(f"⚠️ Error processing letter {letter}: {e}")

    def run(self):
        self.download_reports()
        print("\n✅ Download and logging complete.")
        input("Press Enter to close the browser...")
        self.driver.quit()
        


class CampaignPDFExtractor:
    def __init__(self, input_folder, output_excel):
        self.input_folder = input_folder
        self.output_excel = output_excel

    def detect_form_type(self, page_text):
        if "Full name of contributor" in page_text:
            return "Contribution"
        elif "Payee name" in page_text:
            return "Expenditure"
        return "Unknown"

    def extract_entries(self, blocks, form_type):
        entries = []
        current = {}

        for i, block in enumerate(blocks):
            x0, y0, x1, y1, text, *_ = block
            text = text.strip()

            # Match date
            if re.match(r"\d{1,2}/\d{1,2}/\d{2,4}", text):
                if current:
                    entries.append(current)
                    current = {}
                current['Date'] = text
                continue

            if form_type == "Contribution":
                if "full name of contributor" in text.lower():
                    for j in range(i + 1, len(blocks)):
                        _, y0_next, _, y1_next, text_next, *_ = blocks[j]
                        text_next = text_next.strip()
                        if y0_next > y1 and text_next and not text_next.lower().startswith("out of state pac"):
                            current["Name"] = text_next
                            break

                elif "contributor address" in text.lower():
                    addr = []
                    for j in range(i + 1, len(blocks)):
                        _, y0_next, _, y1_next, text_next, *_ = blocks[j]
                        if any(x in text_next.lower() for x in ["city", "state", "zip"]) or not text_next.strip():
                            break
                        addr.append(text_next.strip())
                    current["Address"] = ' '.join(addr)

                elif "employer" in text.lower():
                    for j in range(i + 1, len(blocks)):
                        _, y0_next, _, y1_next, text_next, *_ = blocks[j]
                        text_next = text_next.strip()
                        if text_next:
                            current["Occupation/Employer"] = text_next
                            break

                elif "amount" in text.lower():
                    if "in-kind" in text.lower():
                        current["Amount"] = "In-kind"
                    else:
                        amt_match = re.search(r"\$?\d[\d,]*(\.\d{2})?", text)
                        if amt_match:
                            current["Amount"] = amt_match.group().replace(",", "").replace("$", "")

            elif form_type == "Expenditure":
                if "payee name" in text.lower():
                    for j in range(i + 1, len(blocks)):
                        _, y0_next, _, y1_next, text_next, *_ = blocks[j]
                        text_next = text_next.strip()
                        if text_next:
                            current["Name"] = text_next
                            break

                elif "payee address" in text.lower():
                    addr = []
                    for j in range(i + 1, len(blocks)):
                        _, y0_next, _, y1_next, text_next, *_ = blocks[j]
                        if any(x in text_next.lower() for x in ['city', 'state', 'zip']) or not text_next.strip():
                            break
                        addr.append(text_next.strip())
                    current["Address"] = ' '.join(addr)

                elif "amount" in text.lower():
                    amt_match = re.search(r"\$?\d[\d,]*(\.\d{2})?", text)
                    if amt_match:
                        current["Amount"] = amt_match.group().replace(",", "").replace("$", "")

        if current:
            entries.append(current)

        for entry in entries:
            entry["Form Type"] = form_type

        return entries

    def extract_data_from_pdf(self, pdf_path, pdf_filename):
        doc = fitz.open(pdf_path)
        all_data = []

        for page in doc:
            if page.number < 3:
                continue

            text = page.get_text()
            blocks = sorted(page.get_text("blocks"), key=lambda b: (b[1], b[0]))  # top to bottom, left to right
            form_type = self.detect_form_type(text)
            if form_type == "Unknown":
                continue

            page_data = self.extract_entries(blocks, form_type)
            for entry in page_data:
                entry["PDF File"] = pdf_filename
            all_data.extend(page_data)

        return all_data

    def process(self):
        all_entries = []

        for filename in os.listdir(self.input_folder):
            if filename.lower().endswith(".pdf"):
                print(f"📄 Processing {filename}...")
                path = os.path.join(self.input_folder, filename)
                all_entries.extend(self.extract_data_from_pdf(path, filename))

        df = pd.DataFrame(all_entries)
        df.to_excel(self.output_excel, index=False)
        print(f"✅ Done! Extracted data saved to: {self.output_excel}")
        
# data_transformer.py


class DataTransformer:
    def __init__(self, input_path, output_path):
        self.input_path = input_path
        self.output_path = output_path

    def extract_name(self, text):
        if pd.isna(text):
            return None
        line = text.split('\n')[0]
        return re.split(r'\d|Amount|In[- ]kind', line)[0].strip()

    def is_date_string(self, s):
        return bool(re.match(r'^\d{1,2}/\d{1,2}/\d{4}', str(s).strip()))

    def clean_amount(self, val):
        try:
            return round(float(val), 2)
        except:
            return None

    def clean_date(self, date_str):
        if pd.isna(date_str):
            return None
        match = re.search(r'(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}[ -](?:jan|feb|...)[a-z]*[ -]\d{2,4})', str(date_str).lower())
        if match:
            try:
                parsed = parser.parse(match.group(1), dayfirst=False, fuzzy=True)
                return parsed.strftime('%m/%d/%Y')
            except:
                return None
        else:
            try:
                parsed = parser.parse(str(date_str), dayfirst=False, fuzzy=True)
                return parsed.strftime('%m/%d/%Y')
            except:
                return None

    def transform(self):
        df = pd.read_excel(self.input_path)
        df['Form Type'] = df['Form Type'].astype(str).str.strip().str.lower()
        df['Name_Original'] = df['Name']
        df['Name'] = df['Name'].apply(self.extract_name)
        df.loc[(df['Form Type'] == 'expenditure') & df['Name_Original'].apply(self.is_date_string), 'Form Type'] = 'expenditure from contribution'

        transformed = []

        for idx, row in df.iterrows():
            form_type = str(row.get("Form Type", "")).strip().lower()
            if form_type == 'skip':
                continue

            name_raw = row.get("Name", "")
            pdf_name = row.get("PDF File", "")
            address_block = str(row.get("Address", "")).strip()
            lines = [line.strip() for line in address_block.split('\n') if line.strip()]
            date = street = city = state = zipcode = amount = ''

            if form_type == "contribution" and len(lines) >= 4:
                date = lines[0]
                street = lines[1]
                city_state_zip = lines[2]
                amount_line = lines[3]

                parts = city_state_zip.split()
                city = parts[0] if len(parts) > 0 else ''
                state = parts[1] if len(parts) > 1 else ''
                zip_match = re.search(r'\b\d{5}\b', city_state_zip)
                zipcode = zip_match.group(0) if zip_match else ''

                amt_match = re.search(r'(\$?\s?\d[\d,]*(?:\.\d{2})?)', amount_line)
                amount = amt_match.group(1).replace('$', '').replace(',', '').strip() if amt_match else ''

            elif form_type == "expenditure" and len(lines) >= 3:
                date = lines[0]
                street = lines[1]
                third_line = lines[2]
                parts = third_line.split(' ', 1)
                amount = re.sub(r'[^\d.]', '', parts[0])
                city_zip = parts[1] if len(parts) > 1 else ''
                zip_match = re.search(r'\b\d{5}\b', city_zip)
                zipcode = zip_match.group(0) if zip_match else ''
                city = city_zip.replace(zipcode, '').strip() if zipcode else city_zip

            elif form_type == "expenditure from contribution":
                if idx + 1 < len(df):
                    next_addr = str(df.at[idx + 1, 'Address']).strip().split()
                    date = row['Name_Original'].split(' ')[0]
                    extracted_name = ' '.join(row['Name_Original'].split(' ')[1:])
                    amount = re.sub(r'[^\d.]', '', next_addr[0]) if next_addr else ''
                    street = ' '.join(next_addr[1:]) if len(next_addr) > 1 else ''
                    zip_match = re.search(r'\b\d{5}\b', street)
                    zipcode = zip_match.group(0) if zip_match else ''
                    city = street.split()[0] if street else ''
                    name_raw = extracted_name
                    df.at[idx + 1, 'Form Type'] = 'skip'
                else:
                    continue
            else:
                continue

            transformed.append({
                "Extracted_Name": name_raw,
                "Date": date,
                "Street_Address": street,
                "City": city,
                "Zipcode": zipcode,
                "Amount": amount,
                "Form Type": form_type.capitalize(),
                "PDF_Name": pdf_name
            })

        result = pd.DataFrame(transformed)
        result = result[result["Extracted_Name"].notna() & (result["Extracted_Name"].str.strip() != "")]
        result = result[result["Amount"].notna() & (result["Amount"].str.strip() != "") & (result["Amount"].str.strip() != ":")]
        result["Amount"] = result["Amount"].apply(self.clean_amount)
        result = result[result["Amount"].notna()]
        result["Date"] = result["Date"].apply(self.clean_date)
        result = result[result["Date"].notna()]

        result.to_excel(self.output_path, index=False)
        print(f"✅ Transformed data saved to {self.output_path}")



# Run the script
if __name__ == "__main__":
    # Step 1: Download PDFs
    downloader = CampaignFinanceDownloader()
    downloader.run()

    # Step 2: Extract data from PDFs to Excel
    extractor = CampaignPDFExtractor(
        input_folder=downloader.download_folder,
        output_excel=os.path.join(downloader.script_directory, "extracted_data.xlsx")
    )
    extractor.process()

    # Step 3: Transform extracted data to cleaned format
    transformer = DataTransformer(
        input_path=os.path.join(downloader.script_directory, "extracted_data.xlsx"),
        output_path=os.path.join(downloader.script_directory, "transformed_data.xlsx")
    )
    transformer.transform()

